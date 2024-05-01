#!/usr/bin/python3
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from lxml import etree
import subprocess
import argparse
import openpyxl
import math
import pandas as pd
import sys
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import PieChart3D, PieChart, ProjectedPieChart, BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
import sqlite3
import pprint
import time
from PIL import Image, ImageDraw, ImageFont 
import argparse
import json
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import csv


parser = argparse.ArgumentParser()
parser.add_argument("--enfoque", "-e", help="Specify the enfoque. Options: cajaGris, cajaNegra", required=True)
parser.parse_args()
args = parser.parse_args()
empresa = 'informe'
enfoqueInforme = args.enfoque # cajaGris cajaNegra


# Add long and short argument
matrizRecomendaciones = []
checklist = []

# Definir los archivos para los informes
#empresas = ["RADICAL","RORIK"]
wbEjecutivo = openpyxl.Workbook() # el informe ejecutivo es igual para todos
wbMatrizVul = openpyxl.Workbook() # Matriz de vulnerabilidades es igual para todos

# Fuentes
Font_MuyAlto = Font(name='Century Gothic', size=12, bold=True, color='FC611C')
Font_Alto = Font(name='Century Gothic', size=12, bold=True, color='FFC000')
Font_Moderado = Font(name='Century Gothic', size=12, bold=True, color='FFFF00')
Font_Bajo = Font(name='Century Gothic', size=12, bold=True, color='00B222')
Font_Info = Font(name='Century Gothic', size=12, bold=True, color='026EBA')


Font_MuyAlto_RORIK = Font(name='Century Gothic', size=10, bold=True, color='FC611C')
Font_Alto_RORIK = Font(name='Century Gothic', size=10, bold=True, color='E36C0A')
Font_Moderado_RORIK = Font(name='Century Gothic', size=10, bold=True, color='FFC000')
Font_Bajo_RORIK = Font(name='Century Gothic', size=10, bold=True, color='026EBA')
Font_muyBajo_RORIK = Font(name='Century Gothic', size=10, bold=True, color='00B222')

italic24Font = Font(size=24, italic=True, bold=True)
Arial10 = Font(name='Century Gothic', size=10)
Arial11Bold = Font(name='Century Gothic', size=11, bold=True)
Arial14Bold = Font(name='Century Gothic', size=14, bold=True)
Arial12BoldWhite = Font(name='Century Gothic', size=12, bold=True, color='FFFFFF')
Arial10Bold = Font(name='Century Gothic', size=10, bold=True)
Calibri10 = Font(name='Century Gothic', size=10)
Calibri10Bold = Font(name='Century Gothic', size=10, bold=True)
Century11Bold = Font(name='Century Gothic', size=11, bold=True)
Calibri12Bold = Font(name='Century Gothic', size=12, bold=True)
Calibri12 = Font(name='Century Gothic', size=12)
Calibri10BoldWhite = Font(name='Century Gothic', size=10, bold=True, color='FFFFFF')
Century12BoldWhite = Font(name='Century Gothic', size=12, bold=True, color='FFFFFF')

Calibri22Bold = Font(name='Century Gothic', size=22, bold=True)

Impact22 = Font(name='Impact', size=22)
Impact36 = Font(name='Impact', size=36)

# Borde
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Rellenos
greyFill = PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
verdeFill = PatternFill(start_color='548DD4', end_color='548DD4', fill_type='solid')

#### RADICAL #####
CRÍTICOFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
altoFill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
medioFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
bajoFill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
brownFill = PatternFill(start_color='EEECE1', end_color='EEECE1', fill_type='solid')
cumpleFill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
noCumpleFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
manualFill = PatternFill(start_color='e2e838', end_color='e2e838', fill_type='solid')
negroFill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
grisFill = PatternFill(start_color='bbbbbb', end_color='bbbbbb', fill_type='solid')

#### RORIK #####
CRÍTICOFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
AltoFill2 = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ModeradoFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
bajoFill2 = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
informativaFill = PatternFill(start_color='95B3D7', end_color='399BE0', fill_type='solid')

############################################################

# Convertir el log en una imagen para el reporte         
#ruta,testType,ip,port,vulType
def log2Image(logPath,imageName):
    #print(f'logPath {logPath} imageName {imageName}')
    # INTERNO/sistemas/logs/vulnerabilidades/192.168.2.247_25_openrelay.txt    
    max_lines_log = 50 #maximo num de lineas del log a convertir en imagen
    logFile = open(logPath)
    vuln_detalles = logFile.read()
    if (vuln_detalles == ''):
        print ("archivo vacio")
        vuln_detalles = "N/A"
    logFile.close()

    vuln_detalles = vuln_detalles.replace("\n\n", "\n") # eliminar saltos de linea excesivos
    vuln_detalles = vuln_detalles.replace("\r", "")  # eliminar     
    vuln_detalles = vuln_detalles.replace("\t", "    ")  # eliminar     
    lista_vuln_detalles = vuln_detalles.split('\n') 
    lineas_log = len(lista_vuln_detalles)
    if lineas_log > max_lines_log:
        lineas_log = max_lines_log

    vuln_detalles = ""    
    i = 0
    #print(f"lista_vuln_detalles {lista_vuln_detalles}")
    imageWidth = 0                    
    for line in lista_vuln_detalles:
        
        if i < max_lines_log :
            vuln_detalles = vuln_detalles + line + "\n"
            #print (f'vuln_detalles {vuln_detalles}')
            i = i + 1
            if (len(line) > imageWidth) and (len(line) < 120): 
               imageWidth = len(line) # determinar cadena con mas longitud (para el ancho de la imagen)
    
    imageWidth = imageWidth * 8
    imageHeight = lineas_log * 15

    #print (f'imageWidth {imageWidth}')
    #print (f'imageHeight {imageHeight}')

    # convertir a imagen
    imagenVulnerabilidad = Image.new('RGB', (imageWidth, imageHeight), color=(255, 255, 255))                    
    canvas = ImageDraw.Draw(imagenVulnerabilidad)
    fnt = ImageFont.truetype("arial.ttf", 12)
    canvas.text((12, 12), vuln_detalles,font=fnt, fill='#000000')
    imagenVulnerabilidad.save(imageName)


root = etree.parse("/usr/share/lanscanner/vulnerabilidades-web.xml").getroot()
total_vulnerabilidades_xml = len(root.getchildren())
#print (f"total_vulnerabilidades_xml en xml {total_vulnerabilidades_xml}")

# Borrar informes anteriores
os.system("rm *.xlsx 2>/dev/null")
os.system("rm estadisticas.csv 2>/dev/null")
os.system("rm *.png 2>/dev/null")



################## CREAR INFORME ############################
#for empresa in empresas:
    
# TODAS LAS PRUEBAS
dominio = ''
segmento = ''

######## #Vulnerabilidades por riesgo ########
total_vul_criticas = 0
total_vul_altas = 0
total_vul_medias = 0
total_vul_bajas = 0
total_vul_info = 0
######

#####################################
                                                                                                    
pruebas_manuales = ['CS-60','CS-10','CS-24','CS-70'] # caja negra
cajaGris = ['CS-02','CS-03','CS-04','CS-05','CS-06','CS-06','CS-09','CS-11','CS-12','CS-13','CS-19','CS-20','CS-21','CS-22','CS-23','CS-25','CS-26','CS-27','CS-36','CS-37','CS-38','CS-50','CS-52','CS-53','CS-54','CS-61','CS-64','CS-65','CS-66','CS-66','CS-67','CS-68']
cajaNegra = ['CS-01','CS-08','CS-39','CS-40','CS-41','CS-42','CS-44','CS-45','CS-46','CS-48','CS-49','CS-56','CS-58','CS-59','CS-60','CS-62','CS-63','CS-69','CS-70']
###### Vulnerabilidades por vectors
total_vuln_externas = 0
total_vuln_internas = 0
#######

resultados_db = '.resultados.db'
print ("\n")
   

# si no existe informe
if not (os.path.isfile(f'{empresa}.xlsx')):
    # crear hoja de calculo
    print (f"Generando informe para {empresa}")
    globals()['wb' + empresa ] = openpyxl.Workbook()


conn = sqlite3.connect(resultados_db)
c = conn.cursor()




# ########################### Vulnerabilidades por criticidad #################

# Vulnerabilidades criticas unicas
vul_criticas = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='passTomcat' or tipo ='zimbraXXE' or tipo ='webshell' or tipo ='backdoorFabrica' or tipo ='ransomware' or tipo ='JoomlaJCKeditor' or tipo ='zerologon' or tipo ='HTTPsys'").fetchone()[0]
total_vul_criticas = total_vul_criticas + vul_criticas
#print(f"total_vul_criticas {total_vul_criticas}")


# Vulnerabilidades altas unicas
vul_altos = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='ms08067' or tipo ='ms17010'  or tipo ='mailPass' or tipo ='defaultPassword' or tipo ='compartidoNFS'  or tipo ='passwordHost' or tipo ='logeoRemoto' or tipo ='heartbleed'  or tipo ='passwordMikroTik' or tipo ='VNCnopass'   or tipo ='perdidaAutenticacion'  or tipo ='slowloris' or tipo ='wordpressPass' or tipo ='conficker' or tipo ='anonymousIPMI' or tipo ='noSQLDatabases' or tipo ='winboxVuln' or tipo ='rmiVuln' or tipo ='SSHBypass' or tipo ='intelVuln' or tipo ='smbrelay' or tipo ='backupWeb' or tipo ='apacheStruts'  or tipo ='IISwebdavVulnerable' or tipo ='shellshock' or tipo ='ciscoASAVuln' or tipo ='SambaCry' or tipo ='misfortune' or tipo ='jbossVuln' or tipo ='passwordBD' or tipo ='wordpressDesactualizado'  or tipo ='poisoning' or tipo ='cipherZeroIPMI'  or tipo ='owaVul' or tipo ='hashRoto' or tipo ='pluginDesactualizado'  or tipo ='github'  or tipo ='Kerberoasting' or tipo ='passwordSFI' or tipo ='restablecimientoPassInseguro' or tipo ='sqli'  or tipo ='openrelayEXTERNO' or tipo ='rdpPass' or tipo ='ddos' or tipo ='passwordAdivinadoServ' or tipo ='passwordAdminWeb' or tipo ='passwordPhpMyadmin' or tipo ='googleRedirect'  or tipo ='phishing' or tipo ='passwordDefecto' or tipo ='BlueKeep'").fetchone()[0]
total_vul_altas = total_vul_altas + vul_altos
#print(f"total_vul_altas {total_vul_altas}")

# Vulnerabilidades medias unicas
vul_medias = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='CS-70' or tipo ='openstreaming'  or tipo ='vulnDahua' or tipo ='directorioLDAP' or tipo ='nullsession' or tipo ='transferenciaDNS' or tipo ='listadoDirectorio' or tipo ='enumeracionUsuarios'  or tipo ='anonymous' or tipo ='ACL' or tipo ='ms12020'  or tipo ='exposicionUsuarios'  or tipo ='upnpAbierto' or tipo ='registroHabilitado' or tipo ='wordpressPingbacks'  or tipo ='scribd' or tipo ='ftpAnonymous' or tipo ='exposicionDatosPersonales'  or tipo ='passwordAdivinadoWin' or tipo ='DoSWeb' or tipo ='openrelayINTERNO'  or tipo ='configuracionInseguraWordpress' or tipo ='VNCbypass' or tipo ='archivosPeligrosos' or tipo ='webdavVulnerable' or tipo ='openWebservice' or tipo ='PrintNightmare' or tipo ='malware' or tipo ='compartidoSMB' or tipo ='gppPassword' or tipo ='escalamientoPrivilegios' or tipo ='noTLS' or tipo ='ghostcat' or tipo ='cmsDesactualizado' or tipo ='spoof'").fetchone()[0]
total_vul_medias = total_vul_medias + vul_medias
#print(f"total_vul_medias {total_vul_medias}")

# Vulnerabilidades bajas unicas
vul_bajas = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='CS-01' or tipo ='CS-51' ").fetchone()[0]
vuln_headers = c.execute("SELECT COUNT ( TIPO)  FROM VULNERABILIDADES WHERE TIPO LIKE 'CS-51-%';").fetchone()[0]
if vuln_headers > 1 :
    vuln_headers = 1
total_vul_bajas = total_vul_bajas + vul_bajas + vuln_headers
#print(f"total_vul_bajas {total_vul_bajas}")

# Vulnerabilidades informativas unicas
vul_info = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='CS-08' or tipo ='CS-49'").fetchone()[0]
total_vul_info = total_vul_info + vul_info

estadisticas = {
    "total_vul_criticas": total_vul_criticas,
    "total_vul_altas": total_vul_altas,
    "total_vul_medias": total_vul_medias,
    "total_vul_bajas": total_vul_bajas,
    "total_vul_info": total_vul_info
}


# Asegurándonos de que todas las entradas del estadisticas sean listas
for key, value in estadisticas.items():
    if not isinstance(value, list):
        estadisticas[key] = value

with open('estadisticas.csv', 'w', newline='') as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=estadisticas.keys())
    writer.writeheader()
    writer.writerow(estadisticas)

# ###################################################




######################  CREAR FICHAS DE VULNERABILIDADES ############
# si no existe informe EXTERNO/INTERNO seteamos a 1
if not (os.path.isfile(f'{empresa}.xlsx')):
    numeroVulnerabilidad = 1

todos_CWE_headers = ''
todas_descripciones_headers = ''
todas_recomendaciones_headers = ''
vuln_header = 0
for i in range(total_vulnerabilidades_xml):
    vulnerabilidadImagen = 0
    conclusion = ""
    recomendacionMatriz = ""
    cod = root[i].find("codVul").text
    nombre = root[i].find("nombre").text 
    criterio_aceptacion = root[i].find("criterio-aceptacion").text 
    criterio_aceptacion = criterio_aceptacion.replace("SALTOLINEA", "\n")
    grupo = root[i].find("grupo").text     
    score = root[i].find("score").text    
    CWE = root[i].find("CWE").text
    CWE = CWE.replace("SALTOLINEA", "\n")
    PCI = root[i].find("PCI").text
    enfoque = root[i].find("enfoque").text
    vector = root[i].find("vector").text 
    riesgoInforme = root[i].find("riesgoInforme").text   
    agente_amenaza = root[i].find("agente_amenaza").text
    impacto_tecnico = root[i].find("impacto_tecnico").text
    impacto_negocio = root[i].find("impacto_negocio").text
    PruebaConcepto = root[i].find("PruebaConcepto").text
    PruebaConcepto = PruebaConcepto.replace("DOMINIOENTIDAD", dominio)
    descripcion = root[i].find("descripcion").text
    recomendacion = root[i].find("recomendacion").text
    recomendacion = recomendacion.replace("DOMINIOENTIDAD", dominio)
    recomendacion = recomendacion.replace("AMPERSAND", '&')
    recomendacion = recomendacion.replace("SALTOLINEA", "\n")
    recomendacionMatriz = root[i].find("recomendacionMatriz").text
    conclusion = root[i].find("conclusion").text
    referenciaweb = root[i].find("referenciaweb").text    
            
    print("cod :'" + cod+"'")

    sql = "SELECT distinct * FROM VULNERABILIDADES WHERE TIPO=\"" + cod + "\";"    
    resSQL = c.execute(sql)
    filas = 1
    hosts = ""     

    #Mix vuln
    if cod in ['CS-51-1','CS-51-2','CS-51-3','CS-51-4']:
        for row in resSQL:
            ip = row[0]
            port = row[1]
            vulType = row[2]
            vuln_detalles = row[3]             
            descripcion = descripcion.replace("HEADER_SEGURO", vuln_detalles)
            
            todas_descripciones_headers = todas_descripciones_headers + descripcion
            todos_CWE_headers = todos_CWE_headers + '\n' + CWE
            todas_recomendaciones_headers = todas_recomendaciones_headers + '\n' + recomendacion

            print(f'vul header: {hosts}  {ip} {port} {vuln_detalles} {todas_descripciones_headers} {todos_CWE_headers} {todas_recomendaciones_headers}')
            hosts = ip +":"+ port + '\n'
            vuln_header = 1

            checklist.append(
            {"ID":'CS-51', 
            'grupo':grupo,
            'Tema':nombre,
            'criterio': criterio_aceptacion, 
            'Estado': 'No cumple'})

        if cod == 'CS-51-4':              
            descripcion = todas_descripciones_headers
            CWE = todos_CWE_headers
            recomendacion = todas_recomendaciones_headers
            cod = 'CS-51'
    
    #print (f'result--> cod {cod} vuln_header {vuln_header}')
    if cod == 'CS-51' and vuln_header == 1:
        #Convertir salida de herramienta a imagen
        if cod in cajaNegra: # si tiene log  
            vulType='CS-49' #solo para log
            logPath = f'logs/vulnerabilidades/{ip}_{port}_{vulType}.txt'
            log2Image(logPath,cod+'.png')     
        filas = filas + 1    
        print (f'filas {filas}')
        print(f'vul headerss: {hosts}  {ip} {port} {vuln_detalles}')        
        
               
            #Caja Negra
    if cod in cajaNegra or cod in pruebas_manuales or cod in cajaGris:
        for row in resSQL:
            ip = row[0]
            port = row[1]
            vulType = row[2]
            
            #Convertir salida de herramienta a imagen
            if cod in cajaNegra: # si tiene log               
                if cod in ['CS-42','CS-08','CS-62']:
                    vulType='responseHeaders' #solo para log
                logPath = f'logs/vulnerabilidades/{ip}_{port}_{vulType}.txt'
                log2Image(logPath,cod+'.png')

            vuln_detalles = row[3]            
            print(f'vul: {hosts}  {ip} {port} {vuln_detalles}')
            hosts = hosts + ip +":"+ port + "\n" + vuln_detalles + "\n"
            filas = filas + 1            
    
    if enfoqueInforme == 'cajaGris' and cod in cajaGris:
        checklist.append(
        {"ID":cod, 
        'grupo':grupo,
        'Tema':nombre,
        'criterio': criterio_aceptacion, 
        'Estado': 'Revision Manual 2'})

    if cod in pruebas_manuales:
        checklist.append(
        {"ID":cod, 
        'grupo':grupo,
        'Tema':nombre,
        'criterio': criterio_aceptacion, 
        'Estado': 'Revision Manual'})


    if filas > 1: 
        print("Generando ficha")   
        #Cargar los datos de las vulnerabilidades a un array de estadisticass
        matrizRecomendaciones.append(
            {"cod":cod, 
            'score':float(score), 
            'vulnerabilidad': nombre, 
            'riesgo': riesgoInforme,                 
            "recomendacionMatriz": recomendacionMatriz, 
            "conclusion": conclusion})
        
        if 'CS-51' not in cod :
            checklist.append(
                {"ID":cod, 
                'grupo':grupo,
                'Tema':nombre,
                'criterio': criterio_aceptacion, 
                'Estado': 'No cumple'})

        globals()['wb' + empresa ].create_sheet()
        sheet = globals()['wb' + empresa]['Sheet']

        sheet.column_dimensions['A'].width = 23
        sheet.column_dimensions['B'].width = 32
        sheet.column_dimensions['C'].width = 27

        sheet.title = str(numeroVulnerabilidad)
        if ("CRÍTICO" in riesgoInforme):
            sheet['A1'].fill = CRÍTICOFill
            sheet['C5'].fill = CRÍTICOFill

        if ("ALTO" in riesgoInforme):
            sheet['A1'].fill = altoFill
            sheet['C5'].fill = altoFill

        if ("MEDIO" in riesgoInforme):
            sheet['A1'].fill = medioFill
            sheet['C5'].fill = medioFill

        if ("BAJO" in riesgoInforme):
            sheet['A1'].fill = bajoFill
            sheet['C5'].fill = bajoFill
        
        if ("INFORMATIVO" in riesgoInforme):
            sheet['A1'].fill = informativaFill
            sheet['C5'].fill = informativaFill

        # Titulo
        sheet.merge_cells('B1:C1')
        sheet.row_dimensions[1].height = 59
        sheet['A1'].font = Impact36
        sheet['B1'].font = Impact22
        sheet['A1'].border = thin_border
        sheet['B1'].border = thin_border
        sheet['C1'].border = thin_border

        sheet['A1'].alignment = Alignment(horizontal="center", vertical='center')
        sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

        sheet['A1'] = str(numeroVulnerabilidad)
        sheet['B1'] = str(nombre).upper()

        
        # grafico
        sheet.row_dimensions[2].height = 76.5
        img = openpyxl.drawing.image.Image('/usr/share/lanscanner/image.png')

        sheet.add_image(img, "D2")
        sheet['A2'].border = thin_border
        sheet['B2'].border = thin_border
        sheet['C2'].border = thin_border
        sheet.merge_cells('A2:C2')

        #imagen de la vulnerabilidad 
        print (f'cod {cod}')
        if cod in cajaNegra :
            imgVul = openpyxl.drawing.image.Image(cod+'.png')
            sheet.add_image(imgVul, "D4")

        # texto del grafico
        sheet.row_dimensions[3].height = 53.2
        sheet['A3'].font = Arial10
        sheet['B3'].font = Arial10
        sheet['C3'].font = Arial10

        sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['C3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

        sheet['A3'].border = thin_border
        sheet['B3'].border = thin_border
        sheet['C3'].border = thin_border

        sheet['A3'] = str(agente_amenaza)
        sheet['B3'] = str(impacto_tecnico)
        sheet['C3'] = str(impacto_negocio)

        # ANALISIS DE RIESGO LABEL
        sheet.merge_cells('A4:C4')
        sheet.row_dimensions[4].height = 29
        sheet['A4'].font = Arial11Bold

        sheet['A4'].border = thin_border
        sheet['B4'].border = thin_border
        sheet['C4'].border = thin_border

        sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A4'] = "ANALISIS DE RIESGO"
        sheet['A4'].fill = greyFill

        # Riesgos
        sheet.row_dimensions[5].height = 34.5
        sheet['A5'].font = Arial10
        sheet['B5'].font = Arial10
        sheet['C5'].font = Arial11Bold

        sheet['A5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['C5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

        sheet['A5'].border = thin_border
        sheet['B5'].border = thin_border
        sheet['C5'].border = thin_border

        sheet['A5'] = "Score: " + str(score)
        sheet['B5'] = "Rol: " + 'Ninguno'
        sheet['C5'] = "RIESGO: " + str(riesgoInforme)

        # vector CVSS
        sheet.row_dimensions[6].height = 20
        sheet.merge_cells('B6:C6')
        sheet['A6'] = "Vector CVSS/4.0:"
        sheet['A6'].border = thin_border
        sheet['B6'].border = thin_border
        sheet['A6'].font = Arial11Bold
        sheet['B6'] =  str(vector)
        sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['B6'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)


        # descripcion LABEL - data
        descriptionHeight = 30 + 20 * (len(descripcion) // 80)
        sheet.row_dimensions[7].height = descriptionHeight
        sheet['A7'].font = Arial11Bold
        sheet['B7'].font = Calibri10

        sheet['A7'].border = thin_border
        sheet['B7'].border = thin_border
        sheet['C7'].border = thin_border

        sheet['A7'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A7'] = "Descripción de la vulnerabilidad:"
        sheet.merge_cells('B7:C7')
        sheet['B7'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
        sheet['B7'] = descripcion

        #Criterio de aceptacion
        sheet['A8'].border = thin_border
        sheet['B8'].border = thin_border
        sheet['C8'].border = thin_border
        sheet.merge_cells('B8:C8')
        sheet['A8'].font = Arial11Bold
        sheet['B8'].font = Arial10
        sheet['A8'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['B8'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)                
        sheet['A8'] = "Criterio de aceptación de seguridad:"
        sheet['B8'] = cod


        # PRUEBA DE CONCEPTO LABEL
        sheet.merge_cells('A9:C9')
        sheet.row_dimensions[9].height = 29
        sheet['A9'].font = Arial11Bold
        sheet['A9'].border = thin_border
        sheet['A9'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A9'] = "PRUEBA DE CONCEPTO"
        sheet['A9'].fill = greyFill


        # HOSTS LABEL - DATA
        hostHeight1 = 20 + 20 * (len(hosts) // 80)
        countNL = hosts.count("SALTOLINEA")
        hostHeight2 = (countNL + 1) * 20
        if hostHeight1 > hostHeight2:
            hostHeight = hostHeight1
        else:
            hostHeight = hostHeight2
        sheet.row_dimensions[10].height = hostHeight
        sheet['A10'].font = Arial11Bold
        sheet['B10'].font = Calibri10

        sheet['A10'].border = thin_border
        sheet['B10'].border = thin_border
        sheet['C10'].border = thin_border

        sheet['A10'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A10'] = "URL afectadas:"

        sheet.merge_cells('B10:C10')
        sheet['B10'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
        hosts = hosts.replace("SALTOLINEA", "\n")
        #print (f' hosts {hosts}')
        sheet['B10'] = ILLEGAL_CHARACTERS_RE.sub(r'', hosts)  

        #condicionesPrevias                
        sheet.merge_cells('B11:C11')
        sheet['A11'].border = thin_border
        sheet['B11'].border = thin_border
        sheet['C11'].border = thin_border
        sheet['A11'].font = Arial11Bold
        sheet['B11'].font = Arial10
        sheet['A11'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['B11'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)                
        sheet['A11'] = "Precondicion:"
        sheet['B11'] = "Ninguna"

        # PRUEBA DE CONCEPTO DATA
        cellHeight = 20 + 20 * (len(PruebaConcepto) // 80)
        sheet.merge_cells('A12:C12')
        sheet.row_dimensions[12].height = cellHeight
        sheet['A12'].font = Arial10

        sheet['A12'].border = thin_border
        sheet['B12'].border = thin_border
        sheet['C12'].border = thin_border

        sheet['A12'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
        PruebaConcepto = PruebaConcepto.replace("SALTOLINEA", "\n")
        sheet['A12'] = PruebaConcepto


        # CONTRAMEDIDAS - LABEL
        sheet.merge_cells('A13:C13')
        sheet.row_dimensions[13].height = 29
        sheet['A13'].font = Arial11Bold

        sheet['A13'].border = thin_border
        sheet['B13'].border = thin_border
        sheet['C13'].border = thin_border

        sheet['A13'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A13'] = "CONTRAMEDIDAS"
        sheet['A13'].fill = greyFill

        # CONTRAMEDIDAS
        sheet.merge_cells('A14:C14')
        cellHeight = 20 + 20 * (len(recomendacion) // 80)

        sheet.row_dimensions[14].height = cellHeight
        sheet['A14'].font = Arial10

        sheet['A14'].border = thin_border
        sheet['B14'].border = thin_border
        sheet['C14'].border = thin_border
        sheet['A14'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
        recomendacion = recomendacion.replace("SALTOLINEA", "\n")
        sheet['A14'] = recomendacion

        # REFERENCIAS - LABEL
        sheet.merge_cells('A15:C15')
        sheet.row_dimensions[15].height = 29
        sheet['A15'].font = Arial11Bold

        sheet['A15'].border = thin_border
        sheet['B15'].border = thin_border
        sheet['C15'].border = thin_border

        sheet['A15'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A15'] = "REFERENCIAS"
        sheet['A15'].fill = greyFill


        # REFERENCIAS CWE
        sheet.merge_cells('A16:B16')            
        sheet['A16'].font = Arial10

        sheet['A16'].border = thin_border
        sheet['B16'].border = thin_border
        sheet['C16'].border = thin_border

        sheet['A16'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
        sheet['A16'] = "Common Weakness Enumeration:"
        sheet['C16'] = CWE

        # REFERENCIAS PCI
        sheet.merge_cells('A17:B17')            
        sheet['A17'].font = Arial10

        sheet['A17'].border = thin_border
        sheet['B17'].border = thin_border
        sheet['C17'].border = thin_border

        sheet['A17'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)            
        sheet['A17'] = "Payment Card Industry (PCI DSS):"
        sheet['C17'] = PCI

        # REFERENCIAS WEB
        sheet.merge_cells('A18:C18')
        cellHeight = 20 + 20 * (len(referenciaweb) // 80)
        sheet.row_dimensions[18].height = cellHeight
        sheet['A18'].font = Arial10

        sheet['A18'].border = thin_border
        sheet['B18'].border = thin_border
        sheet['C18'].border = thin_border

        sheet['A18'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
        referenciaweb = referenciaweb.replace("SALTOLINEA", "\n")
        referenciaweb = referenciaweb.replace("TAB", "\t")
        sheet['A18'] = referenciaweb

        numeroVulnerabilidad = numeroVulnerabilidad + 1
    else:
        if cod not in pruebas_manuales and 'CS-51' not in cod and cod not in cajaGris:
            checklist.append(
                {"ID":cod, 
                'grupo':grupo,
                'Tema':nombre, 
                'criterio': criterio_aceptacion, 
                'Estado': 'Cumple'})          

                                 
##### CHECK LIST ##########
print("Generando checklist")
globals()['wb' + empresa].create_sheet()
sheet = globals()['wb' + empresa]['Sheet']
sheet.title = 'checklist'  # Change title.

i = 1                
sheet.column_dimensions['A'].width = 8
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 45
sheet.column_dimensions['D'].width = 10

current_group = ''
i = 0 # elementos checklist
index = 1 # tuplas excel
while i < len(checklist):
    #print (f'i={i} index={index}')
    tupla = checklist[i]
    ID = tupla["ID"]
    grupo = tupla["grupo"]
    Tema = tupla["Tema"]
    criterio = tupla["criterio"]    
    Estado = tupla["Estado"]

    if (current_group != grupo): #nuevo grupo --> set headers                          
        sheet.merge_cells('A'+str(index)+':D'+str(index))
        sheet['A' + str(index)].fill = negroFill   
        sheet['A' + str(index)].font = Century12BoldWhite    
        sheet['A' + str(index)] = grupo
        sheet['A' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        index = index + 1

        sheet['A' + str(index)] = "ID"
        sheet['B' + str(index)] = "Tema"
        sheet['C' + str(index)] = "Criterio de aceptación"
        sheet['D' + str(index)] = "Estado"
        sheet['A' + str(index)].font = Century11Bold
        sheet['B' + str(index)].font = Century11Bold
        sheet['C' + str(index)].font = Century11Bold
        sheet['D' + str(index)].font = Century11Bold
        sheet['A' + str(index)].fill = grisFill  
        sheet['B' + str(index)].fill = grisFill
        sheet['C' + str(index)].fill = grisFill
        sheet['D' + str(index)].fill = grisFill
          
        sheet['A' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['B' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['C' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['D' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A' + str(index)].border = thin_border
        sheet['B' + str(index)].border = thin_border
        sheet['C' + str(index)].border = thin_border
        sheet['D' + str(index)].border = thin_border
        index = index + 1
        current_group=grupo        
    else: # tupla normal        
        sheet['A' + str(index)] = ID       
        sheet['B' + str(index)] = Tema
        sheet['C' + str(index)] = criterio
        sheet['D' + str(index)] = Estado
        
        sheet['A' + str(index)].alignment = Alignment(horizontal="center", vertical='center',wrap_text=True)
        sheet['B' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['C' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['D' + str(index)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        sheet['A' + str(index)].border = thin_border
        sheet['B' + str(index)].border = thin_border
        sheet['C' + str(index)].border = thin_border
        sheet['D' + str(index)].border = thin_border

        if ("No cumple" in Estado):
            sheet['D' + str(index)].fill = noCumpleFill
        else:
            sheet['D' + str(index)].fill = cumpleFill

        if ("Revision Manual" in Estado):
            sheet['D' + str(index)].fill = manualFill

        index = index + 1
        i = i + 1

##### CONCLUSIONES Y RECOMENDACIONES ####            
print("Generando conclusiones y recomendaciones")
globals()['wb' + empresa].create_sheet()
sheet = globals()['wb' + empresa]['Sheet']
sheet.title = 'Conclusiones y recomendaciones'  # Change title.

#ordenar conclusiones
matrizRecomendaciones = sorted(matrizRecomendaciones, key=lambda k: k['score'],reverse=True) 

i = 1                
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 18
sheet.column_dimensions['C'].width = 23
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 32
sheet['A' + str(i)] = "Nº"
sheet['B' + str(i)] = "ID"
sheet['C' + str(i)] = "Vulnerabilidad identificada"
sheet['D' + str(i)] = "Nivel de riesgo"
sheet['E' + str(i)] = "Contramedida "
sheet['A' + str(i)].font = Calibri12Bold
sheet['B' + str(i)].font = Calibri12Bold
sheet['C' + str(i)].font = Calibri12Bold
sheet['D' + str(i)].font = Calibri12Bold
sheet['E' + str(i)].font = Calibri12Bold
sheet['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['E' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A' + str(i)].border = thin_border
sheet['B' + str(i)].border = thin_border
sheet['C' + str(i)].border = thin_border
sheet['D' + str(i)].border = thin_border
sheet['E' + str(i)].border = thin_border

#ordenar recomendaciones
matrizRecomendaciones = sorted(matrizRecomendaciones, key=lambda k: k['score'],reverse=True) 
for tupla in matrizRecomendaciones:        
    cod = tupla["cod"]
    vulnerabilidad = tupla["vulnerabilidad"]
    riesgo = tupla["riesgo"]
    conclusion = tupla["conclusion"]        
    recomendacionMatriz = tupla["recomendacionMatriz"]        
    #vectorInforme = tupla["vectorInforme"]        
    sheet['A' + str(i+1)] = i        
    sheet['B' + str(i+1)] = cod
    sheet['C' + str(i+1)] = vulnerabilidad
    sheet['D' + str(i+1)] = riesgo
    sheet['E' + str(i+1)] = recomendacionMatriz
    sheet['A' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center',wrap_text=True)
    sheet['B' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['C' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['D' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['E' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A' + str(i + 1)].border = thin_border
    sheet['B' + str(i + 1)].border = thin_border
    sheet['C' + str(i + 1)].border = thin_border
    sheet['D' + str(i + 1)].border = thin_border
    sheet['E' + str(i + 1)].border = thin_border

    if ("CRÍTICO" in riesgo):
        sheet['D' + str(i+1)].fill = CRÍTICOFill

    if ("ALTO" in riesgo):
        sheet['D' + str(i+1)].fill = altoFill

    if ("MEDIO" in riesgo):
        sheet['D' + str(i+1)].fill = medioFill

    if ("BAJO" in riesgo):
        sheet['D' + str(i+1)].fill = bajoFill

    if ("INFORMATIVO" in riesgo):
        sheet['D' + str(i+1)].fill = informativaFill
    i = i + 1

print (f'Escribiendo en  archivo {empresa}.xlsx')        
globals()['wb' + empresa].save(f'{empresa}.xlsx')

